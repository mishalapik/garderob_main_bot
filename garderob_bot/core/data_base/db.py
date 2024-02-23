import psycopg2
from aiogram import Bot
from aiogram.types import Message
from openpyxl import load_workbook
from aiogram.types import FSInputFile
from core.config.data import bot_god,elite_ops,operators

conn = psycopg2.connect(database='garderob_main_db',user='postgres',host='localhost',port='5432',password='yesterday')
cur = conn.cursor()


def clear_db(a:str,b:int):
    fn = 'core/data_base/dbgarderobbot.xlsx'
    wb = load_workbook(fn)
    ws = wb[a]
    ws.delete_cols(1, b)
    wb.save(fn)
    wb.close()


async def data_base_copy(message:Message,bot:Bot):
    if message.from_user.id in elite_ops:
        clear_db('клиенты',4)
        clear_db('заказы', 8)
        fn = 'core/data_base/dbgarderobbot.xlsx'
        wb = load_workbook(fn)
        ws = wb['клиенты']
        await message.answer('<b>сборка базы данных началась...</b>')


        try:
            global cur, conn
            cur.execute('SELECT COUNT(*) FROM customer_info')
        except psycopg2.ProgrammingError as exc:
            conn.rollback()
            cur.execute('SELECT COUNT(*) FROM customer_info')
        except psycopg2.InterfaceError as exc:
            conn = psycopg2.connect(database='garderob_main_db',user='postgres',host='localhost',port='5432',password='yesterday')
            cur = conn.cursor()
            cur.execute('SELECT COUNT(*) FROM customer_info')
        except psycopg2.OperationalError:
            conn = psycopg2.connect(database='garderob_main_db', user='postgres', host='localhost', port='5432',
                                    password='yesterday')
            cur = conn.cursor()
            cur.execute('SELECT COUNT(*) FROM customer_info')
        except psycopg2.errors.InFailedSqlTransaction:
            conn = psycopg2.connect(database='garderob_main_db', user='postgres', host='localhost', port='5432',
                                    password='yesterday')
            cur = conn.cursor()
            cur.execute('SELECT COUNT(*) FROM customer_info')
        l =  int(cur.fetchone()[0])
        cur.execute('SELECT (customer_id,customer_fio,phone_number,mail) FROM customer_info')
        ws.append(['ID','ФИО','Номер телефона','Почта'])
        for i in range(l):
            a = cur.fetchone()[0].split(',')
            ws.append([str(a[0])[1:],a[1][1:-1],a[2],str(a[3])[1:-2]])
        cur.execute('SELECT COUNT(*) FROM orders')
        l = int(cur.fetchone()[0])
        cur.execute('SELECT (customer_id,order_id,product_name,product_size,product_price,product_discount,order_date,adress) FROM orders')
        ws = wb['заказы']
        ws.append(['tgID','ID Заказа','Название товара','Размер','Цена','Скидка','Дата','Адрес'])
        for i in range(l):
            a = cur.fetchone()[0].split(',')
            ws.append([str(a[0])[1:],a[1],a[2],a[3],a[4],a[5],a[6],str(a[7])[:-1]])
        wb.save(fn)
        wb.close()
        base = FSInputFile(fn)
        await bot.send_document(chat_id=message.from_user.id,document=base)


async def clear_data_base(message:Message):
    if message.from_user.id in elite_ops:
        try:
            global cur, conn
            cur.execute('TRUNCATE TABLE order_state CASCADE')
            cur.execute('TRUNCATE TABLE customer_info CASCADE')
            cur.execute('TRUNCATE TABLE orders CASCADE')

        except psycopg2.ProgrammingError as exc:
            conn.rollback()
            cur.execute('TRUNCATE TABLE order_state CASCADE')
            cur.execute('TRUNCATE TABLE customer_info CASCADE')
            cur.execute('TRUNCATE TABLE orders CASCADE')
        except psycopg2.InterfaceError as exc:
            conn = psycopg2.connect(database='garderob_main_db',user='postgres',host='localhost',port='5432',password='yesterday')
            cur = conn.cursor()
            cur.execute('TRUNCATE TABLE order_state CASCADE')
            cur.execute('TRUNCATE TABLE customer_info CASCADE')
            cur.execute('TRUNCATE TABLE orders CASCADE')
        except psycopg2.OperationalError:
            conn = psycopg2.connect(database='garderob_main_db', user='postgres', host='localhost', port='5432',
                                    password='yesterday')
            cur = conn.cursor()
            cur.execute('TRUNCATE TABLE order_state CASCADE')
            cur.execute('TRUNCATE TABLE customer_info CASCADE')
            cur.execute('TRUNCATE TABLE orders CASCADE')
        except psycopg2.errors.InFailedSqlTransaction:
            conn = psycopg2.connect(database='garderob_main_db', user='postgres', host='localhost', port='5432',
                                    password='yesterday')
            cur = conn.cursor()
            cur.execute('TRUNCATE TABLE order_state CASCADE')
            cur.execute('TRUNCATE TABLE customer_info CASCADE')
            cur.execute('TRUNCATE TABLE orders CASCADE')
        await message.answer('база данных очищена')



async def show_gods(message:Message):
    if message.from_user.id in bot_god:
        msg = ''
        for i in range(len(bot_god)):
            msg+=f'{bot_god[i]}\n'
        await message.answer(msg)

async def add_operator(message:Message):
    if message.from_user.id in bot_god:
        operators.append(int(message.text.split()[1]))
async def remove_operator(message:Message):
    if message.from_user.id in bot_god:
        operators.remove(int(message.text.split()[1]))
async def show_ops(message:Message):
    if message.from_user.id in bot_god:
        msg = ''
        for i in range(len(operators)):
            msg += f'{operators[i]}\n'
        await message.answer(msg)

async def add_el_op(message:Message):
    if message.from_user.id in bot_god:
        elite_ops.append(int(message.text.split()[1]))
async def remove_el_op(message:Message):
    if message.from_user.id in bot_god:
        elite_ops.remove(int(message.text.split()[1]))
async def show_el_ops(message:Message):
    if message.from_user.id in bot_god:
        msg = ''
        for i in range(len(elite_ops)):
            msg += f'{elite_ops[i]}\n'
        await message.answer(msg)

